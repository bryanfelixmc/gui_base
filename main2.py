import sys
import pandas as pd
from openpyxl import load_workbook
from pyautocad import Autocad, APoint
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication


class Information:
    """Clase para almacenar información"""
    def __init__(self):
        self.data = {}  # Use a dictionary to store key-value pairs
    def set_value(self, key, value):
        """Set a value for a given key"""
        
        self.data[key] = value
        print(f"-> set \"{self.data}\"")
    def get_value(self, key):
        """Get a value for a given key"""
        print(f"-> get \"{self.data}\"")
        return self.data.get(key, "")  # Return an empty string if the key does not exist

class WindowManager:
    """Clase para administrar la transicion entre ventanas"""
    def __init__(self, info):
        self.info = info
        self.current_window = None

    def show_window(self, window_class):
        """Mostrar una nueva ventana y ocultar la actual"""
        if self.current_window is not None:
            self.current_window.hide()  # Hide the current window
        self.current_window = window_class(self.info)  # Create a new window instance
        self.current_window.show()  # Show the new window

    def setup_comboBox(self, comboBox, elementoslista, key, defaultindex):
        """Configurar comboBox con los elementos dados y almacenar el valor seleccionado"""
        # Rellenar valores de comboBox
        comboBox.addItems(elementoslista)
        
        # Recuperar valores almacenados
        value = self.info.get_value(key)
        index = elementoslista.index(value) if value in elementoslista else defaultindex
        
        # Obtener el index actual y almacenar el default value
        comboBox.setCurrentIndex(index)
        self.info.set_value(key, elementoslista[index] if value == "" else value)
        
        # Conecte la señal currentTextChanged para actualizar la clase Information
        comboBox.currentTextChanged.connect(lambda x: self.info.set_value(key, x))   
    
    def setup_lineEdit(self, lineEdit, key):
        """Setup a lineEdit and store its value."""
        default_value = self.info.get_value(key)  # Recuperar el valor almacenado
        lineEdit.setText(default_value)  # Establezca el texto actual de QLineEdit en el valor almacenado
        
        # Conecte la señal textChanged para almacenar el valor de QLineEdit
        lineEdit.textChanged.connect(lambda x: self.info.set_value(key, x))


class Ventana1(QDialog):
    def __init__(self, info):
        super(Ventana1, self).__init__()
        loadUi("ventana1.ui", self)
        self.info = info  # Store the information instance
        self.next_button.clicked.connect(lambda: window_manager.show_window(Ventana2))  # Connect the button to the method

        # Configuracion de los QComboBox
        window_manager.setup_comboBox(
            comboBox=self.comboBox_2,
            elementoslista=["Muy ligero: 12.7 mm/kV", "Ligero: 16 mm/kV", "Medio: 20 mm/kV", "Pesado: 25 mm/kV", "Muy pesado: 31 mm/kV"],
            key="NIV_DE_CONTAM",
            defaultindex=4  # Default to "Muy pesado"
        )
        # Configuracion de los QLineEdit
        window_manager.setup_lineEdit(
            lineEdit=self.lineEdit,
            key="V_NOM"  # Setup lineEdit with key
        )
        window_manager.setup_lineEdit(
            lineEdit=self.lineEdit_2,
            key="V_MAX"  # Setup lineEdit with key
        )        
        window_manager.setup_lineEdit(
            lineEdit=self.lineEdit_3,
            key="HZ"  # Setup lineEdit with key
        )
        window_manager.setup_lineEdit(
            lineEdit=self.lineEdit_4,
            key="ICC"  # Setup lineEdit with key
        )


class Ventana2(QDialog):
    def __init__(self, info):
        super(Ventana2, self).__init__()
        loadUi("ventana2.ui", self)
        self.info = info  # Store the information instance
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana1)) # Connect the back button to the method
        self.next_button.clicked.connect(lambda: window_manager.show_window(Ventana3))  # Connect the next button to a new method
        self.pushButton_linebay.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1(QDialog):
    def __init__(self, info):
        super(Ventana2_1, self).__init__()
        loadUi("ventana2_1.ui", self)
        self.info = info  # Store the information instance
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2)) # Connect the back button to the method
        self.pushButton_TT.clicked.connect(lambda: window_manager.show_window(Ventana2_1_1))
        self.pushButton_TC.clicked.connect(lambda: window_manager.show_window(Ventana2_1_2))
        self.pushButton_SL.clicked.connect(lambda: window_manager.show_window(Ventana2_1_3))
        self.pushButton_SB.clicked.connect(lambda: window_manager.show_window(Ventana2_1_4))
        self.pushButton_IN.clicked.connect(lambda: window_manager.show_window(Ventana2_1_5))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_1(QDialog):
    def __init__(self, info):
        super(Ventana2_1_1, self).__init__()
        loadUi("ventana2_1_1.ui", self)
        self.info = info  # Store the information instance
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_2(QDialog):
    def __init__(self, info):
        super(Ventana2_1_2, self).__init__()
        loadUi("ventana2_1_2.ui", self)
        self.info = info  # Store the information instance
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_3(QDialog):
    def __init__(self, info):
        super(Ventana2_1_3, self).__init__()
        loadUi("ventana2_1_3.ui", self)
        self.info = info  # Store the information instance
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_4(QDialog):
    def __init__(self, info):
        super(Ventana2_1_4, self).__init__()
        loadUi("ventana2_1_4.ui", self)
        self.info = info  # Store the information instance
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_5(QDialog):
    def __init__(self, info):
        super(Ventana2_1_5, self).__init__()
        loadUi("ventana2_1_5.ui", self)
        self.info = info  # Store the information instance
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana
class Ventana3(QDialog):
    def __init__(self, info):
        super(Ventana3, self).__init__()
        loadUi("ventana3.ui", self)
        self.info = info  # Store the information instance
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2))  # Connect the back button to the method
        self.next_button.clicked.connect(lambda: window_manager.show_window(Ventana4))  # Connect the next button to a new method
        self.pushButton.clicked.connect(self.draw_bahia)  # Connect the draw button to the method

    def draw_bahia(self):
        """Create an instance of Bahia and draw in AutoCAD."""
        ruta = "BD.xlsx" # Get the Excel file path from the info instance
        nombre_hoja = "bahia"  # Replace with the actual sheet name
        p0 = [0, 10]  # Example offset for drawing in AutoCAD
        # Create an instance of the Bahia class
        bahia_instance = Bahia(ruta, nombre_hoja, p0)
        print("Drawing Bahia elements in AutoCAD...")
        bahia_instance.draw_bahia_in_autocad()  # Call the method to draw in AutoCAD

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana



class Ventana4(QDialog):
    def __init__(self, info):
        super(Ventana4, self).__init__()
        loadUi("ventana4.ui", self)
        self.info = info  # Store the information instance
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana3))  # Connect the back button to the method


    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana




#INICIO PROGRAMA DE MANUEL

class DataProcessor:
    def __init__(self, input_file):
        self.input_file = input_file
        self.df_bahia = None
        self.df_metrado = None
        self.tags = {
            "DESCARGADORES DE SOBRETENSIÓN": "PR",
            "TRAMPA DE ONDA": "TO",
            "TRANSFORMADOR DE TENSIÓN CAPACITIVO": "TTC",
            "SECCIONADOR DE LÍNEA": "SL",
            "TRANSFORMADOR DE CORRIENTE": "TC",
            "INTERRUPTOR DE POTENCIA": "IN",
            "SECCIONADOR DE BARRA": "SB"
        }

    def read_data(self):
        self.df_bahia = pd.read_excel(self.input_file, sheet_name="bahia")
        self.df_metrado = pd.read_excel(self.input_file, sheet_name="metrado")
        self.clean_data()

    def clean_data(self):
        self.df_bahia["Elemento"] = self.df_bahia["Elemento"].str.strip().str.upper()
        self.df_metrado["DESCRIPCION"] = self.df_metrado["DESCRIPCION"].str.upper()

    def generate_bahia_metrado(self):
        rows_bahia = []
        item_n = 0

        for i, row in self.df_metrado.iterrows():
            descripcion = row["DESCRIPCION"]
            cantidad = row["CANTIDAD"]

            if descripcion.startswith("BAHÍA DE"):
                item_n += 1
                item = f"{item_n}"
                rows_bahia.append({"ITEM": item, "DESCRIPCION": descripcion.title(), "CANTIDAD": cantidad})
                subitem = 1
            else:
                item = f"{item_n}.{subitem}"
                rows_bahia.append({"ITEM": item, "DESCRIPCION": descripcion.title(), "CANTIDAD": cantidad})
                self.add_bahia_details(rows_bahia, descripcion)
                subitem += 1

        return pd.DataFrame(rows_bahia)

    def add_bahia_details(self, rows_bahia, descripcion):
        tag_base = self.tags.get(descripcion.upper(), "")
        if tag_base:
            tag_rows = self.df_bahia[self.df_bahia["Elemento"].str.startswith(tag_base)]
            tag_rows = tag_rows.drop_duplicates(subset=["Descripción", "Valor", "Unidad"])
            for _, det in tag_rows.iterrows():
                if str(det["Descripción"]).strip().upper() != "N.A.":
                    detalle = f"    {det['Descripción']} = {det['Valor']} {det['Unidad']}"
                    rows_bahia.append({"ITEM": "", "DESCRIPCION": detalle, "CANTIDAD": ""})

    def generate_total_metrado(self):
        df_total = self.df_metrado[~self.df_metrado["DESCRIPCION"].str.startswith("BAHÍA DE")].copy()
        df_total = df_total.groupby("DESCRIPCION", as_index=False)["CANTIDAD"].sum().sort_values(by="DESCRIPCION")

        rows_total = []
        usados = set()

        for i, row in enumerate(df_total.itertuples(), start=1):
            descripcion = row.DESCRIPCION.title()
            cantidad = row.CANTIDAD
            item = f"{i}"
            rows_total.append({"ITEM": item, "EQUIPO": descripcion, "CANTIDAD": cantidad})
            self.add_total_details(rows_total, row.DESCRIPCION, usados)

        return pd.DataFrame(rows_total)

    def add_total_details(self, rows_total, descripcion, usados):
        tag_base = self.tags.get(descripcion.upper(), "")
        if tag_base and tag_base not in usados:
            usados.add(tag_base)
            tag_rows = self.df_bahia[self.df_bahia["Elemento"].str.startswith(tag_base)]
            tag_rows = tag_rows.drop_duplicates(subset=["Descripción", "Valor", "Unidad"])
            for _, det in tag_rows.iterrows():
                if str(det["Descripción"]).strip().upper() != "N.A.":
                    detalle = f"    {det['Descripción']} = {det['Valor']} {det['Unidad']}"
                    rows_total.append({"ITEM": "", "EQUIPO": detalle, "CANTIDAD": ""})

    def export_to_excel(self, df_bahia_metrado, df_total_metrado, output_file):
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df_bahia_metrado.to_excel(writer, sheet_name="Metrado por bahía", index=False)
            df_total_metrado.to_excel(writer, sheet_name="Metrado total por equipo", index=False)

    def process(self, output_file):
        self.read_data()
        df_bahia_metrado = self.generate_bahia_metrado()
        df_total_metrado = self.generate_total_metrado()
        self.export_to_excel(df_bahia_metrado, df_total_metrado, output_file)
        print(f"Metrado generado como '{output_file}'")

#FIN PROGRAMA DE MANUEL


#INICIO PROGRAMA DE JOEL


class ExcelDataHandler:
    def __init__(self, ruta, nombre_hoja):
        self.wb = load_workbook(ruta, data_only=True)
        self.ws = self.wb[nombre_hoja]
        self.dict_df = self.load_dataframes()

    def load_dataframes(self):
        """Load data from Excel tables into a dictionary of DataFrames."""
        dict_df = {}
        for tbl_name, tbl_ref in self.ws.tables.items():
            cells = self.ws[tbl_ref]
            headers = [c.value for c in cells[0]]
            data_rows = [[c.value for c in row] for row in cells[1:]]
            dict_df[tbl_name] = pd.DataFrame(data_rows, columns=headers)
        return dict_df

    def get_text_from_dict(self):
        """Extract text from the DataFrame based on specific conditions."""
        text = ""
        for df in self.dict_df.values():
            sub_df = df[df['Elemento'].duplicated(keep=False)]
            for index, row in sub_df.iterrows():
                text += f"{row['Descripción']} {row['Propiedad']} : {row['Valor']} {row['Unidad']}\n"
        return text.strip()


class AutoCADHandler:
    def __init__(self):
        self.acad = Autocad(create_if_not_exists=True, visible=True)
        self.doc = self.acad.ActiveDocument
        self.lib_objetos = self.doc.blocks
        self.listado = [elem.name for elem in self.lib_objetos]

    def draw_elements(self, elementos, p0):
        """Draw elements in AutoCAD based on user-defined parameters."""
        defined_width = 80
        separacion_derecha_eje = 30
        for i, instance in enumerate(elementos):
            block_name = instance.get("block_name")
            if block_name in self.listado:
                self.acad.model.InsertBlock(APoint(instance.get("x") + p0[0], instance.get("y"), 0), block_name, 1, 1, 1, 0)
                self.acad.model.AddMText(APoint(instance.get("x") + separacion_derecha_eje + p0[0], instance.get("y") + 3, 0), defined_width, block_name)
                self.acad.model.AddMText(APoint(instance.get("x") + separacion_derecha_eje + p0[0], instance.get("y") - (i + 1) * 3, 0), defined_width, instance.get("Descripcion"))

    def clear_autocad(self):

        for obj in list(self.doc.ModelSpace):
            obj.Delete()
        print("Cleaning & Ploting again...")


class Bahia(ExcelDataHandler, AutoCADHandler):
    def __init__(self, ruta, nombre_hoja, p0):
        ExcelDataHandler.__init__(self, ruta, nombre_hoja)
        AutoCADHandler.__init__(self)
        self.p0 = p0
        self.elementos = self.initialize_elementos()
        self.input_usuario()

    def initialize_elementos(self):
        """Initialize the elements with default values."""
        return [
            {"x": 0, "y": 200, "Descripcion": "", "block_name": ""},
            {"x": 0, "y": 170, "Descripcion": "", "block_name": ""},
            {"x": 0, "y": 140, "Descripcion": "", "block_name": ""},
            {"x": 0, "y": 110, "Descripcion": "", "block_name": ""},
            {"x": 0, "y": 85, "Descripcion": "", "block_name": ""},
            {"x": 0, "y": 50, "Descripcion": "", "block_name": ""},
            {"x": 0, "y": 20, "Descripcion": "", "block_name": ""}
        ]

    def input_usuario(self):
        """Collect user input for various elements."""
        self.elementos[0]["block_name"] = "SALIDA_LINEA"
        self.elementos[1]["block_name"] = "Pararrayos c-cd1"
        self.elementos[1]["Descripcion"] = self.get_text_from_dict()
        '''
        self.elegir_tt()
        self.elegir_tc()
        '''
        self.elementos[3]["block_name"] = "SL_AT_AV"

        self.elementos[5]["block_name"] = "IP_AT"
        self.elementos[6]["block_name"] = "SB_AT_AC"
    '''
    def elegir_tt(self):
        """Select the type of transformer based on user input."""
        options = {
            "1": "TTI_AT_2S",
            "2": "TTC_AT_1S",
            "3": "TTC_AT_2S",
            "4": "TTC_AT_3S"
        }
        while True:
            val = input("\nElija el tipo de transformador de tension:\n1) Inductivo 2 dev.\n2) Capacitivo 1 dev.\n3) Capacitivo 2 dev.\n4) Capacitivo 3 dev.\n--> ")
            if val in options:
                self.elementos[2]["block_name"] = options[val]
                break
            else:
                print("Eliga la opcion nuevamente")

    def elegir_tc(self):
        """Select the type of current transformer based on user input."""
        options = {
            "1": "CT_AT_3S",
            "2": "CT_AT_4S"
        }
        while True:
            val = input("\nElija el tipo de transformador de corriente:\n1) tres devanados\n2) cuatro devanados\n--> ")
            if val in options:
                self.elementos[4]["block_name"] = options[val]
                break
            else:
                print("Eliga la opcion nuevamente")
    '''
    def draw_bahia_in_autocad(self):
        """Draw Bahia elements in AutoCAD."""
        self.draw_elements(self.elementos, self.p0)

    def clear_autocad(self):
        """Clear AutoCAD model space."""
        self.clear_autocad()



#FIN PROGRAMA DE JOEL

def main():
    app = QApplication(sys.argv)
    info = Information()  # Create an instance of the Information class
    global window_manager  # Declare window_manager as global to access it in other classes
    window_manager = WindowManager(info)  # Create an instance of the WindowManager
    window_manager.show_window(Ventana1)  # Show the first window

    #INICIO PROGRAMA MANUEL
    processor = DataProcessor("BD.xlsx")
    processor.process("Metrado_Final.xlsx")
    #FIN PROGRAMA MANUEL
    '''
    #INICIO PROGRAMA JOEL
    bahia1 = Bahia("BD.xlsx", "bahia", p0=[0, 0])
    bahia1.clear_autocad()  # Clear AutoCAD before drawing
    bahia1.draw_bahia_in_autocad()
    #FIN PROGRAMA JOEL
    '''

    try:
        sys.exit(app.exec_())
    except SystemExit:
        print("Exiting")

if __name__ == "__main__":
    main()