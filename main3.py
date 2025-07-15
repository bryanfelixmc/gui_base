import sys
import pandas as pd
from typing import List,Dict, Any
from openpyxl import load_workbook
from pyautocad import Autocad, APoint
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication

class Information:
    """Clase para almacenar información"""
    def __init__(self):
        self.data = {}  # Use a dictionary to store key-value pairs
    def set_value(self, key, value):
        """Set a value for a given key"""
        
        self.data[key] = value
        print(f"-> set \"{self.data}\"")
    def get_value(self, key):
        """Get a value for a given key"""
        print(f"-> get \"{self.data}\"")
        return self.data.get(key, "")  # Return an empty string if the key does not exist

class WindowManager:
    """Clase para administrar la transicion entre ventanas"""
    def __init__(self, info: Information):
        self.info = info
        self.current_window = None

    def show_window(self, window_class):
        """Mostrar una nueva ventana y ocultar la actual"""
        if self.current_window is not None:
            self.current_window.hide()  # Hide the current window
        self.current_window = window_class()  # Create a new window instance
        self.current_window.show()  # Show the new window

    def setup_comboBox(self, comboBox, elementoslista, key, defaultindex):
        """Configurar comboBox con los elementos dados y almacenar el valor seleccionado"""
        # Rellenar valores de comboBox
        comboBox.addItems(elementoslista)
        
        # Recuperar valores almacenados
        value = self.info.get_value(key)
        index = elementoslista.index(value) if value in elementoslista else defaultindex
        
        # Obtener el index actual y almacenar el default value
        comboBox.setCurrentIndex(index)
        self.info.set_value(key, elementoslista[index] if value == "" else value)
        
        # Conecte la señal currentTextChanged para actualizar la clase Information
        comboBox.currentTextChanged.connect(lambda x: self.info.set_value(key, x))   
    
    def setup_lineEdit(self, lineEdit, key):
        """Setup a lineEdit and store its value."""
        default_value = self.info.get_value(key)  # Recuperar el valor almacenado
        lineEdit.setText(default_value)  # Establezca el texto actual de QLineEdit en el valor almacenado
        
        # Conecte la señal textChanged para almacenar el valor de QLineEdit
        lineEdit.textChanged.connect(lambda x: self.info.set_value(key, x))


    def setup_groupBox(self, groupBox, key):
        """Configurar un QGroupBox con botones de opción y almacenar el valor seleccionado"""
        # Recuperar el valor almacenado
        selected_value = self.info.get_value(key)

        # Iterar sobre los botones de opción en el groupBox
        for radioButton in groupBox.findChildren(QtWidgets.QRadioButton):
            # Si el texto del botón de opción coincide con el valor almacenado, marcarlo como seleccionado
            if radioButton.text() == selected_value:
                radioButton.setChecked(True)
            else:
                radioButton.setChecked(False)

        # Conectar la señal toggled para actualizar la clase Information
        for radioButton in groupBox.findChildren(QtWidgets.QRadioButton):
            radioButton.toggled.connect(lambda checked, btn=radioButton: self.info.set_value(key, btn.text()) if checked else None)



class Ventana1(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana1.ui", self)
        self.next_button.clicked.connect(lambda: window_manager.show_window(Ventana2))  # Connect the button to the method

        # Configuracion de los QComboBox
        window_manager.setup_comboBox(
            comboBox=self.comboBox_2,
            elementoslista=["Muy ligero: 12.7 mm/kV", "Ligero: 16 mm/kV", "Medio: 20 mm/kV", "Pesado: 25 mm/kV", "Muy pesado: 31 mm/kV"],
            key="NIV_DE_CONTAM",
            defaultindex=4  # Default to "Muy pesado"
        )
        # Configuracion de los QLineEdit
        window_manager.setup_lineEdit(
            lineEdit=self.lineEdit,
            key="V_NOM"  # Setup lineEdit with key
        )
        window_manager.setup_lineEdit(
            lineEdit=self.lineEdit_2,
            key="V_MAX"  # Setup lineEdit with key
        )        
        window_manager.setup_lineEdit(
            lineEdit=self.lineEdit_3,
            key="HZ"  # Setup lineEdit with key
        )
        window_manager.setup_lineEdit(
            lineEdit=self.lineEdit_4,
            key="ICC"  # Setup lineEdit with key
        )
        # Configuracion de los QGroupBox
        window_manager.setup_groupBox(
            groupBox=self.groupBox,
            key= "SE_CONFIG"
        )
        window_manager.setup_groupBox(
            groupBox=self.groupBox_2,
            key= "SE_TECN"
        )

class Ventana2(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana2.ui", self)
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana1)) # Connect the back button to the method
        self.next_button.clicked.connect(lambda: window_manager.show_window(Ventana3))  # Connect the next button to a new method
        self.pushButton_linebay.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana2_1.ui", self)
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2)) # Connect the back button to the method
        self.pushButton_TT.clicked.connect(lambda: window_manager.show_window(Ventana2_1_1))
        self.pushButton_TC.clicked.connect(lambda: window_manager.show_window(Ventana2_1_2))
        self.pushButton_SL.clicked.connect(lambda: window_manager.show_window(Ventana2_1_3))
        self.pushButton_SB.clicked.connect(lambda: window_manager.show_window(Ventana2_1_4))
        self.pushButton_IN.clicked.connect(lambda: window_manager.show_window(Ventana2_1_5))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_1(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana2_1_1.ui", self)
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_2(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana2_1_2.ui", self)
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_3(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana2_1_3.ui", self)
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_4(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana2_1_4.ui", self)
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana2_1_5(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana2_1_5.ui", self)
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2_1))

    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana
class Ventana3(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana3.ui", self)
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana2))  # Connect the back button to the method
        self.next_button.clicked.connect(lambda: window_manager.show_window(Ventana4))  # Connect the next button to a new method
        self.pushButton.clicked.connect(self.draw_bahia)  # Connect the draw button to the method

    def draw_bahia(self):
        pass
    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

class Ventana4(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("ventana4.ui", self)
        self.prev_button.clicked.connect(lambda: window_manager.show_window(Ventana3))  # Connect the back button to the method


    def go_to_ventana(self,ventana):
        window_manager.show_window(ventana)  # Use the window manager to show ventana

#INICIO PROGRAMA DE MANUEL

class DataProcessor:
    def __init__(self, input_file):
        self.input_file = input_file
        self.df_bahia = None
        self.df_metrado = None
        self.tags = {
            "DESCARGADORES DE SOBRETENSIÓN": "PR",
            "TRAMPA DE ONDA": "TO",
            "TRANSFORMADOR DE TENSIÓN CAPACITIVO": "TTC",
            "SECCIONADOR DE LÍNEA": "SL",
            "TRANSFORMADOR DE CORRIENTE": "TC",
            "INTERRUPTOR DE POTENCIA": "IN",
            "SECCIONADOR DE BARRA": "SB"
        }

    def read_data(self):
        self.df_bahia = pd.read_excel(self.input_file, sheet_name="bahia")
        self.df_metrado = pd.read_excel(self.input_file, sheet_name="metrado")
        self.clean_data()

    def clean_data(self):
        self.df_bahia["Elemento"] = self.df_bahia["Elemento"].str.strip().str.upper()
        self.df_metrado["DESCRIPCION"] = self.df_metrado["DESCRIPCION"].str.upper()

    def generate_bahia_metrado(self):
        rows_bahia = []
        item_n = 0

        for i, row in self.df_metrado.iterrows():
            descripcion = row["DESCRIPCION"]
            cantidad = row["CANTIDAD"]

            if descripcion.startswith("BAHÍA DE"):
                item_n += 1
                item = f"{item_n}"
                rows_bahia.append({"ITEM": item, "DESCRIPCION": descripcion.title(), "CANTIDAD": cantidad})
                subitem = 1
            else:
                item = f"{item_n}.{subitem}"
                rows_bahia.append({"ITEM": item, "DESCRIPCION": descripcion.title(), "CANTIDAD": cantidad})
                self.add_bahia_details(rows_bahia, descripcion)
                subitem += 1

        return pd.DataFrame(rows_bahia)

    def add_bahia_details(self, rows_bahia, descripcion):
        tag_base = self.tags.get(descripcion.upper(), "")
        if tag_base:
            tag_rows = self.df_bahia[self.df_bahia["Elemento"].str.startswith(tag_base)]
            tag_rows = tag_rows.drop_duplicates(subset=["Descripción", "Valor", "Unidad"])
            for _, det in tag_rows.iterrows():
                if str(det["Descripción"]).strip().upper() != "N.A.":
                    detalle = f"    {det['Descripción']} = {det['Valor']} {det['Unidad']}"
                    rows_bahia.append({"ITEM": "", "DESCRIPCION": detalle, "CANTIDAD": ""})

    def generate_total_metrado(self):
        df_total = self.df_metrado[~self.df_metrado["DESCRIPCION"].str.startswith("BAHÍA DE")].copy()
        df_total = df_total.groupby("DESCRIPCION", as_index=False)["CANTIDAD"].sum().sort_values(by="DESCRIPCION")

        rows_total = []
        usados = set()

        for i, row in enumerate(df_total.itertuples(), start=1):
            descripcion = row.DESCRIPCION.title()
            cantidad = row.CANTIDAD
            item = f"{i}"
            rows_total.append({"ITEM": item, "EQUIPO": descripcion, "CANTIDAD": cantidad})
            self.add_total_details(rows_total, row.DESCRIPCION, usados)

        return pd.DataFrame(rows_total)

    def add_total_details(self, rows_total, descripcion, usados):
        tag_base = self.tags.get(descripcion.upper(), "")
        if tag_base and tag_base not in usados:
            usados.add(tag_base)
            tag_rows = self.df_bahia[self.df_bahia["Elemento"].str.startswith(tag_base)]
            tag_rows = tag_rows.drop_duplicates(subset=["Descripción", "Valor", "Unidad"])
            for _, det in tag_rows.iterrows():
                if str(det["Descripción"]).strip().upper() != "N.A.":
                    detalle = f"    {det['Descripción']} = {det['Valor']} {det['Unidad']}"
                    rows_total.append({"ITEM": "", "EQUIPO": detalle, "CANTIDAD": ""})

#FIN PROGRAMA DE MANUEL


class ExcelDataHandler:
    def __init__(self, ruta, nombre_hoja):
        self.wb = load_workbook(ruta, data_only=True)
        self.ws = self.wb[nombre_hoja]
        self.dict_df = self.load_dataframes()

    def load_dataframes(self):
        """Load data from Excel tables into a dictionary of DataFrames."""
        dict_df = {}
        for tbl_name, tbl_ref in self.ws.tables.items():
            cells = self.ws[tbl_ref]
            headers = [c.value for c in cells[0]]
            data_rows = [[c.value for c in row] for row in cells[1:]]
            dict_df[tbl_name] = pd.DataFrame(data_rows, columns=headers)
        return dict_df

    def get_text_from_dict(self):
        """Extract text from the DataFrame based on specific conditions."""
        text = ""
        for df in self.dict_df.values():
            sub_df = df[df['Elemento'].duplicated(keep=False)]
            for index, row in sub_df.iterrows():
                text += f"{row['Descripción']} {row['Propiedad']} : {row['Valor']} {row['Unidad']}\n"
        return text.strip()

class AutoCADHandler:
    def __init__(self):
        self.acad = Autocad(create_if_not_exists=True, visible=True)
        self.doc = self.acad.ActiveDocument
        self.lib_objetos = self.doc.blocks
        self.listado = [elem.name for elem in self.lib_objetos]

    def draw_elements(self, elementos, p0):
        """Draw elements in AutoCAD based on user-defined parameters."""
        defined_width = 80
        separacion_derecha_eje = 30
        for i, instance in enumerate(elementos):
            block_name = instance.get("block_name")
            if block_name in self.listado:
                self.acad.model.InsertBlock(APoint(instance.get("x") + p0[0], instance.get("y"), 0), block_name, 1, 1, 1, 0)
                self.acad.model.AddMText(APoint(instance.get("x") + separacion_derecha_eje + p0[0], instance.get("y") + 3, 0), defined_width, block_name)
                self.acad.model.AddMText(APoint(instance.get("x") + separacion_derecha_eje + p0[0], instance.get("y") - (i + 1) * 3, 0), defined_width, instance.get("Descripcion"))

    def clear_autocad(self):

        for obj in list(self.doc.ModelSpace):
            obj.Delete()
        print("Cleaning & Ploting again...")


#Construyendo los objetos de Bahia

class Pararrayo:
    def __init__(self, nom_aux: str):
        self.nom_aux = nom_aux  # Nombre del pararrayo
        self.Ur = 150  # kV, Tensión asignada
        self.Uc = 151  # kV, Tensión continua de operación
        self.Cl = 1  # Clase
        self.In = 2  # kA, Corriente de descarga nominal
        self.Is = 40  # kA, Corriente asignada de cortocircuito
        self.t = 0.5  # s, Duración de corriente de cortocircuito
        self.Wth = 12  # kJ/kV, Energía térmica
        self.material = "ZnO"  # Material de construcción
        self.acad_block_name="Pararrayos c-cd1" #Nombre del bloque para AutoCad
        self.acad_x=0 #posicion x para AutoCad
        self.acad_y=230 #posicion y para AutoCad
    
    def set(self, Ur: float = None, Uc: float = None, Cl: int = None,
            In: float = None, Is: float = None, t: float = None,
            Wth: float = None, material: str = None, acad_block_name: str = None,
            acad_x: str = None, acad_y: str = None) -> None:
        if Ur is not None:
            self.Ur = Ur
        if Uc is not None:
            self.Uc = Uc
        if Cl is not None:
            self.Cl = Cl
        if In is not None:
            self.In = In
        if Is is not None:
            self.Is = Is
        if t is not None:
            self.t = t
        if Wth is not None:
            self.Wth = Wth
        if material is not None:
            self.material = material
        if acad_block_name is not None:
            self.acad_block_name = acad_block_name
        if acad_x is not None:
            self.acad_x = acad_x
        if acad_y is not None:
            self.acad_y = acad_y

class TrampaDeOnda:
    def __init__(self, nom_aux: str):
        self.nom_aux = nom_aux  # Nombre de la trampa de onda
        self.LtN = 0.4  # mH, Inductancia asignada
        self.Um = 220  # kV, Tensión máxima del sistema
        self.Ir = 2000  # A, Corriente asignada de servicio continuo
        self.Ik = 31.5  # kA, Corriente de corta duración de servicio asignada
        self.RB = 400  # ohm, Resistencia de sintonía
        self.ancho_banda_inferior = "[50-100] kHz"  # Ancho de banda del dispositivo de sintonía (límite inferior)
        self.ancho_banda_superior = "[100-400] kHz"  # Ancho de banda del dispositivo de sintonía (límite superior)
        self.acad_block_name="TO_AT" #Nombre del bloque para AutoCad
        self.acad_x=0 #posicion x para AutoCad
        self.acad_y=200 #posicion y para AutoCad

    def set(self, LtN: float = None, Um: float = None, Ir: float = None,
            Ik: float = None, RB: float = None,
            ancho_banda_inferior: str = None, ancho_banda_superior: str = None,
              acad_block_name: str = None, acad_x: str = None, acad_y: str = None) -> None:
        if LtN is not None:
            self.LtN = LtN
        if Um is not None:
            self.Um = Um
        if Ir is not None:
            self.Ir = Ir
        if Ik is not None:
            self.Ik = Ik
        if RB is not None:
            self.RB = RB
        if ancho_banda_inferior is not None:
            self.ancho_banda_inferior = ancho_banda_inferior
        if ancho_banda_superior is not None:
            self.ancho_banda_superior = ancho_banda_superior
        if acad_block_name is not None:
            self.acad_block_name = acad_block_name
        if acad_x is not None:
            self.acad_x = acad_x
        if acad_y is not None:
            self.acad_y = acad_y

class TT:
    def __init__(self, nom_aux: str):
        self.nom_aux = nom_aux  # Nombre del transformador
        self.Ur = 245  # kV, Tensión asignada
        self.Ud = 460  # kVr.m.s., Tensión soportada asignada de corta duración a frecuencia industrial
        self.Usys = 220  # kV, Tensión del sistema
        self.Up = 1050  # kVp, Nivel básico de aislamiento
        self.Upr = "220/\u221A3"  # kV, Relación primaria
        self.USr1 = "0.11 /\u221A3"  # kV, Relación del secundario
        self.USr2 = "0.11 /\u221A3"  # kV, Relación del secundario
        self.Sr = 20  # VA, Potencia de salida asignada
        self.clase_medicion = 0.2  # Clase de precisión de medición
        self.clase_proteccion = "3P"  # Clase de precisión de protección
        self.acad_block_name="TTC_AT_1S" #Nombre del bloque para AutoCad
        self.acad_x=0 #posicion x para AutoCad
        self.acad_y=170 #posicion y para AutoCad

    def set(self, Ur: float = None, Ud: float = None, Usys: float = None,
            Up: float = None, Upr: float = None, USr1: float = None,
            USr2: float = None, Sr: float = None,
            clase_medicion: float = None, clase_proteccion: str = None,
              acad_block_name: str = None, acad_x: str = None, acad_y: str = None) -> None:
        if Ur is not None:
            self.Ur = Ur
        if Ud is not None:
            self.Ud = Ud
        if Usys is not None:
            self.Usys = Usys
        if Up is not None:
            self.Up = Up
        if Upr is not None:
            self.Upr = Upr
        if USr1 is not None:
            self.USr1 = USr1
        if USr2 is not None:
            self.USr2 = USr2
        if Sr is not None:
            self.Sr = Sr
        if clase_medicion is not None:
            self.clase_medicion = clase_medicion
        if clase_proteccion is not None:
            self.clase_proteccion = clase_proteccion
        if acad_block_name is not None:
            self.acad_block_name = acad_block_name
        if acad_x is not None:
            self.acad_x = acad_x
        if acad_y is not None:
            self.acad_y = acad_y

class SeccLinea:
    def __init__(self, nom_aux: str):
        self.nom_aux = nom_aux  # Nombre del seccionador de línea
        self.Ur = 245  # kV, Tensión asignada
        self.Ud = 460  # kVr.m.s., Tensión soportada asignada de corta duración a frecuencia industrial
        self.Up = 1050  # kVp, Nivel básico de aislamiento
        self.Ir = 2000  # A, Corriente asignada en servicio continuo
        self.Ik = 40  # kA, Corriente de corta duración admisible asignada
        self.t = 0.5  # s, Duración de corriente de cortocircuito
        self.accionamiento = "Tripolar"  # Accionamiento
        self.tipo_constructivo = "Apertura central"  # Tipo constructivo
        self.cuchilla_tierra = "Si"  # Cuchilla de puesta a tierra
        self.acad_block_name="SL_AT_AV" #Nombre del bloque para AutoCad
        self.acad_x=0 #posicion x para AutoCad
        self.acad_y=140 #posicion y para AutoCad

    def set(self, Ur: float = None, Ud: float = None, Up: float = None,
            Ir: float = None, Ik: float = None, t: float = None,
            accionamiento: str = None, tipo_constructivo: str = None,
            cuchilla_tierra: str = None, acad_block_name: str = None, 
            acad_x: str = None, acad_y: str = None) -> None:
        if Ur is not None:
            self.Ur = Ur
        if Ud is not None:
            self.Ud = Ud
        if Up is not None:
            self.Up = Up
        if Ir is not None:
            self.Ir = Ir
        if Ik is not None:
            self.Ik = Ik
        if t is not None:
            self.t = t
        if accionamiento is not None:
            self.accionamiento = accionamiento
        if tipo_constructivo is not None:
            self.tipo_constructivo = tipo_constructivo
        if cuchilla_tierra is not None:
            self.cuchilla_tierra = cuchilla_tierra
        if acad_block_name is not None:
            self.acad_block_name = acad_block_name
        if acad_x is not None:
            self.acad_x = acad_x
        if acad_y is not None:
            self.acad_y = acad_y

class TC:
    def __init__(self, nom_aux: str):
        self.nom_aux = nom_aux  # Nombre del transformador de corriente
        self.Ur = 245  # kV, Tensión asignada
        self.Ipr = "500-1000"  # A, Valores normalizados de corriente primaria asignada
        self.Isr = 1  # A, Valores normalizados de corriente secundaria asignada
        self.cantidad_nucleos_medicion = 1  # Cantidad de núcleos de medición
        self.cantidad_nucleos_proteccion = 3  # Cantidad de núcleos de protección
        self.clase_medicion = 0.2  # Clase de precisión de medición
        self.clase_proteccion = "5P20"  # Clase de precisión de protección
        self.Ith = 40  # kA, Corriente de cortocircuito térmica asignada
        self.t = 0.5  # s, Duración de corriente de cortocircuito
        self.Sr = 20  # VA, Potencia de salida asignada
        self.acad_block_name="CT_AT_4S" #Nombre del bloque para AutoCad
        self.acad_x=0 #posicion x para AutoCad
        self.acad_y=110 #posicion y para AutoCad

    def set(self, Ur: float = None, Ipr: str = None, Isr: float = None,
            cantidad_nucleos_medicion: int = None, cantidad_nucleos_proteccion: int = None,
            clase_medicion: float = None, clase_proteccion: str = None,
            Ith: float = None, t: float = None, Sr: float = None, acad_block_name: str = None, 
            acad_x: str = None, acad_y: str = None) -> None:
        if Ur is not None:
            self.Ur = Ur
        if Ipr is not None:
            self.Ipr = Ipr
        if Isr is not None:
            self.Isr = Isr
        if cantidad_nucleos_medicion is not None:
            self.cantidad_nucleos_medicion = cantidad_nucleos_medicion
        if cantidad_nucleos_proteccion is not None:
            self.cantidad_nucleos_proteccion = cantidad_nucleos_proteccion
        if clase_medicion is not None:
            self.clase_medicion = clase_medicion
        if clase_proteccion is not None:
            self.clase_proteccion = clase_proteccion
        if Ith is not None:
            self.Ith = Ith
        if t is not None:
            self.t = t
        if Sr is not None:
            self.Sr = Sr
        if acad_block_name is not None:
            self.acad_block_name = acad_block_name
        if acad_x is not None:
            self.acad_x = acad_x
        if acad_y is not None:
            self.acad_y = acad_y
class Interruptor:
    def __init__(self,nom_aux,bahia_aux):
        self.nom_aux=nom_aux
        self.bahia_aux=bahia_aux
        self.codigo = "IN-XXXX"  # Código del interruptor
        self.tension_asignada = 245  # kV
        self.tension_soportada_fi = 460  # kVr.m.s.
        self.nivel_basico_aislamiento = 1050  # kVp
        self.corriente_asignada_servicio_continuo = 2000  # A
        self.corriente_corta_duracion = 40  # kA
        self.duracion_corriente_cortocircuito = 0.5  # s
        self.accionamiento = "UNI-TRIPOLAR"  # Tipo de accionamiento
        self.diseño_externo = "Tanque vivo"  # Diseño externo
        self.mecanismo_operacion = "Resortes"  # Mecanismo de operación
        self.medio_interrupcion = "SF6"  # Medio de interrupción
        self.factor_primer_polo = 1.5  # Factor de primer polo
        self.numero_camaras_interrupcion_por_polo = 1  # Número de cámaras de interrupción por polo
        self.acad_block_name="IP_AT" #Nombre del bloque para AutoCad
        self.acad_x=0 #posicion x para AutoCad
        self.acad_y=85 #posicion y para AutoCad
    
    def set(self, 
            tension_asignada: float = None, 
            tension_soportada_fi: float = None,
            nivel_basico_aislamiento: float = None,
            corriente_asignada_servicio_continuo: float = None,
            corriente_corta_duracion: float = None,
            duracion_corriente_cortocircuito: float = None,
            accionamiento: str = None,
            diseño_externo: str = None,
            mecanismo_operacion: str = None,
            medio_interrupcion: str = None,
            factor_primer_polo: float = None,
            numero_camaras_interrupcion_por_polo: int = None, acad_block_name: str = None, 
            acad_x: str = None, acad_y: str = None) -> None:

        if tension_asignada is not None:
            self.tension_asignada = tension_asignada
        if tension_soportada_fi is not None:
            self.tension_soportada_fi = tension_soportada_fi
        if nivel_basico_aislamiento is not None:
            self.nivel_basico_aislamiento = nivel_basico_aislamiento
        if corriente_asignada_servicio_continuo is not None:
            self.corriente_asignada_servicio_continuo = corriente_asignada_servicio_continuo
        if corriente_corta_duracion is not None:
            self.corriente_corta_duracion = corriente_corta_duracion
        if duracion_corriente_cortocircuito is not None:
            self.duracion_corriente_cortocircuito = duracion_corriente_cortocircuito
        if accionamiento is not None:
            self.accionamiento = accionamiento
        if diseño_externo is not None:
            self.diseño_externo = diseño_externo
        if mecanismo_operacion is not None:
            self.mecanismo_operacion = mecanismo_operacion
        if medio_interrupcion is not None:
            self.medio_interrupcion = medio_interrupcion
        if factor_primer_polo is not None:
            self.factor_primer_polo = factor_primer_polo
        if numero_camaras_interrupcion_por_polo is not None:
            self.numero_camaras_interrupcion_por_polo = numero_camaras_interrupcion_por_polo
        if acad_block_name is not None:
            self.acad_block_name = acad_block_name
        if acad_x is not None:
            self.acad_x = acad_x
        if acad_y is not None:
            self.acad_y = acad_y


class SeccBarra:
    def __init__(self, nom_aux: str):
        self.nom_aux = nom_aux  # Nombre del seccionador de barra
        self.Ur = 245  # kV, Tensión asignada
        self.Ud = 460  # kVr.m.s., Tensión soportada asignada de corta duración a frecuencia industrial
        self.Up = 1050  # kVp, Nivel básico de aislamiento
        self.Ir = 2000  # A, Corriente asignada en servicio continuo
        self.Ik = 40  # kA, Corriente de corta duración admisible asignada
        self.t = 0.5  # s, Duración de corriente de cortocircuito
        self.accionamiento = "Tripolar"  # Accionamiento
        self.tipo_constructivo = "Apertura central"  # Tipo constructivo
        self.cuchilla_tierra = "NO"  # Cuchilla de puesta a tierra
        self.acad_block_name="SB_AT_AC" #Nombre del bloque para AutoCad
        self.acad_x=0 #posicion x para AutoCad
        self.acad_y=50 #posicion y para AutoCad

    def set(self, Ur: float = None, Ud: float = None, Up: float = None,
            Ir: float = None, Ik: float = None, t: float = None,
            accionamiento: str = None, tipo_constructivo: str = None,
            cuchilla_tierra: str = None, acad_block_name: str = None, 
            acad_x: str = None, acad_y: str = None) -> None:
        if Ur is not None:
            self.Ur = Ur
        if Ud is not None:
            self.Ud = Ud
        if Up is not None:
            self.Up = Up
        if Ir is not None:
            self.Ir = Ir
        if Ik is not None:
            self.Ik = Ik
        if t is not None:
            self.t = t
        if accionamiento is not None:
            self.accionamiento = accionamiento
        if tipo_constructivo is not None:
            self.tipo_constructivo = tipo_constructivo
        if cuchilla_tierra is not None:
            self.cuchilla_tierra = cuchilla_tierra
        if acad_block_name is not None:
            self.acad_block_name = acad_block_name
        if acad_x is not None:
            self.acad_x = acad_x
        if acad_y is not None:
            self.acad_y = acad_y

class Bahia:   
    def __init__(self, nom_aux: str, tipo_aux: str):
        self.nom_aux = nom_aux  # Nombre de la bahía (ej. "bay1", "bay2")
        self.tipo_aux = tipo_aux  # Tipo de bahía (línea, acople, transformación, medición, etc.)
        self.equipos: Dict[str, Any] = {}  # Diccionario para almacenar equipos, clave: nombre/código del equipo, valor: objeto del equipo
    def agregar_equipo(self, equipo_cualquiera: Any) -> None:
        if hasattr(equipo_cualquiera, 'nom_aux'):
            self.equipos[equipo_cualquiera.nom_aux] = equipo_cualquiera
        else:
            raise ValueError("El equipo debe tener un atributo 'nom_aux' para ser agregado.")
    def obtener_equipo(self, nombre_equipo: str) -> Any:
        return self.equipos.get(nombre_equipo, None)


def main():
    app = QApplication(sys.argv)
    info = Information()  # Create an instance of the Information class
    global window_manager  # Declare window_manager as global to access it in other classes
    window_manager = WindowManager(info)  # Create an instance of the WindowManager
    window_manager.show_window(Ventana1)  # Show the first window


    try:
        sys.exit(app.exec_())
    except SystemExit:
        print("Exiting")

if __name__ == "__main__":
    main()